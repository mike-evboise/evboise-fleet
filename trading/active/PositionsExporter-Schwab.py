#!/usr/bin/env python3
"""
Positions Exporter - Schwab
----------------------------------------------------------------------------

Purpose:
    This script connects to the Schwab API to export option and equity positions
    from the user’s Schwab brokerage account, enriches them with live market data,
    and writes the final result into an Excel workbook.

Overall Flow:
    1. Load configuration and environment variables from `.env`
    2. Refresh OAuth2 token using saved Schwab credentials
    3. Retrieve all account positions via the Schwab Trading API
    4. Write raw position data into Excel (Phase 1)
    5. Fetch live option chain and quote data from Schwab Market Data API
    6. Update each position row with computed and fetched metrics (Phase 2)
    7. Save enriched workbook to disk with runtime progress output and colors

Storage Locations:
    • Token JSON → C:\\Users\\mnc35\\evboise-fleet\\_dev\\Working_Files\\OAuth\\schwab_token.json
    • Excel export → C:\\Users\\mnc35\\evboise-fleet\\_dev\\Working_Files\\positions.xlsx
    • Raw API dump → C:\\Users\\mnc35\\evboise-fleet\\_dev\\Working_Files\\accounts.json

Dependencies:
    - requests        : for HTTP API calls
    - openpyxl        : for Excel export
    - dotenv          : for environment variable loading
    - typing, pathlib : for type safety and filesystem paths

-------------------------------------------------------------------------------
"""

import os, json, base64, time, requests
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------
# Initialization
# --------------------------------------------------------------------------
load_dotenv()

CONFIG = {
    "TOKEN_PATH": Path(r"C:\Users\mnc35\evboise-fleet\_dev\Working_Files\OAuth\schwab_token.json"),
    "TOKEN_URL": "https://api.schwabapi.com/v1/oauth/token",
    "ACCT_URL": "https://api.schwabapi.com/trader/v1/accounts",
    "CHAINS_URL": "https://api.schwabapi.com/marketdata/v1/chains",
    "QUOTES_URL": "https://api.schwabapi.com/marketdata/v1/quotes",
    "XLSX_FILE": Path(r"C:\Users\mnc35\evboise-fleet\_dev\Working_Files\positions.xlsx"),
    "RAW_FILE": Path(r"C:\Users\mnc35\evboise-fleet\_dev\Working_Files\accounts.json"),
}
MAP = {"BRKB": "BRK.B", "BRKA": "BRK.A"}

# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def first(*v):
    return next((x for x in v if x is not None), None)

def parse(sym: str) -> Optional[Tuple[str, str, str, float]]:
    if not isinstance(sym, str) or len(sym) < 21:
        return None
    root, yymmdd, cp = MAP.get(sym[:6].strip(), sym[:6].strip()), sym[6:12], sym[12].upper()
    try:
        strike = int(sym[13:].lstrip('0') or '0') / 1000
        expiry = f"{2000 + int(yymmdd[:2])}-{yymmdd[2:4]}-{yymmdd[4:]}"
        return root, expiry, cp, strike
    except:
        return None

# --------------------------------------------------------------------------
# OAuth
# --------------------------------------------------------------------------
class SchwabAuth:
    def __init__(self):
        self.cid, self.csec, self.redirect = [os.getenv(k) for k in
            ("SCHWAB_CLIENT_ID", "SCHWAB_CLIENT_SECRET", "SCHWAB_REDIRECT_URI")]
        if not all((self.cid, self.csec, self.redirect)):
            raise SystemExit("❌ Missing .env credentials")

    def refresh(self) -> str:
        p = CONFIG["TOKEN_PATH"]
        t = json.loads(p.read_text()) if p.exists() else {}
        h = {
            "Authorization": "Basic " + base64.b64encode(f"{self.cid}:{self.csec}".encode()).decode(),
            "Content-Type": "application/x-www-form-urlencoded",
        }
        r = requests.post(CONFIG["TOKEN_URL"],
            data={"grant_type": "refresh_token", "refresh_token": t.get("refresh_token"), "redirect_uri": self.redirect},
            headers=h, timeout=30)
        r.raise_for_status()
        j = r.json()
        p.write_text(json.dumps(j, indent=2))
        return j["access_token"]

# --------------------------------------------------------------------------
# Market Data
# --------------------------------------------------------------------------
class OptionFetcher:
    def __init__(self, t: str):
        self.t = t

    def get(self, u: str, p: Dict[str, Any]) -> Dict[str, Any]:
        r = requests.get(u, headers={"Authorization": f"Bearer {self.t}"}, params=p, timeout=30)
        return r.json() if r.ok else {}

    def chain(self, s, e, cp, k):
        p = {
            "symbol": s,
            "contractType": "CALL" if cp == "C" else "PUT",
            "strategy": "SINGLE",
            "strike": str(int(k)) if float(k).is_integer() else f"{k:.2f}",
            "fromDate": e,
            "toDate": e,
            "includeQuotes": "TRUE"
        }
        for alt in (s, {"BRK.B": "BRK/B", "BRKB": "BRK-B"}.get(s)):
            if not alt:
                continue
            p["symbol"] = alt
            j = self.get(CONFIG["CHAINS_URL"], p)
            if j:
                return j
        return {}

    def extract(self, j: Dict[str, Any], cp, e, k, fdata=None) -> Dict[str, Any]:
        em = j.get("callExpDateMap" if cp == "C" else "putExpDateMap", {})
        if not em:
            return {}
        ex = next((x for x in em if x.split(":")[0] == e), None)
        if not ex or not em[ex]:
            return {}
        sk = next((x for x in em[ex] if abs(float(x) - k) < 1e-6), None)
        if not sk:
            return {}
        eobj = em[ex][sk][0]
        base = {f: eobj.get(f) for f in (
            "delta","theta","volatility","totalVolume","openInterest","timeValue",
            "highPrice","lowPrice","closePrice","theoreticalVolatility","daysToExpiration")}
        u = first(*[(j.get("underlying", {}) or {}).get(x) for x in ("mark","last","close","price")]) or j.get("underlyingPrice") or 0
        d = eobj.get("optionDeliverablesList")
        du = d[0].get("deliverableUnits") if isinstance(d, list) and d else None
        return {
            **base,
            "deliverableUnits": du,
            "UnderlyingPrice": u,
            "DivYield": (fdata or {}).get("divYield") or 0,
            "DivAmount": (fdata or {}).get("divAmount") or 0,
            "DivExDate": (fdata or {}).get("divExDate") or None,
            "LastEarningsDate": (fdata or {}).get("lastEarningsDate") or None,
            "NextDivExDate": (fdata or {}).get("nextDivExDate") or None,
            "DTE": eobj.get("daysToExpiration") or 0
        }

# --------------------------------------------------------------------------
# Excel Export
# --------------------------------------------------------------------------
class Exporter:
    HDRS = ["accountId","symbol","cusip","description","assetType","longQuantity","shortQuantity","netQuantity",
            "averagePrice","marketValue","maintenanceRequirement","averageLongPrice","longOpenProfitLoss",
            "shortOpenProfitLoss","netOpenProfitLoss","costBasis","currentDayProfitLoss","currentDayProfitLossPct",
            "delta","theta","volatility","totalVolume","openInterest","timeValue","highPrice","lowPrice","closePrice",
            "deliverableUnits","theoreticalVolatility","UnderlyingPrice","DivYield","DivAmount","DivExDate",
            "LastEarningsDate","NextDivExDate","DTE"]

    def phase1(self, a: List[Dict[str, Any]]):
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Workbook.active returned None")
        ws = cast(Worksheet, ws)
        ws.title = "Positions"
        ws.append(self.HDRS)

        for x in a:
            s = x.get("securitiesAccount", {})
            if not isinstance(s, dict):
                continue
            aid = s.get("accountNumber", "UNKNOWN")
            for p in s.get("positions", []):
                i = p.get("instrument", {})
                at = i.get("assetType")
                lq, sq, avg = [float(p.get(k) or 0) for k in ("longQuantity", "shortQuantity", "averagePrice")]
                cb = (lq + sq) * avg * (100 if at == "OPTION" else 1)
                ws.append([
                    aid, i.get("symbol"), i.get("cusip"), i.get("description"), at, lq, sq, lq - sq, avg,
                    p.get("marketValue"), p.get("maintenanceRequirement"), p.get("averageLongPrice"),
                    p.get("longOpenProfitLoss") or 0, p.get("shortOpenProfitLoss") or 0,
                    (p.get("longOpenProfitLoss") or 0) + (p.get("shortOpenProfitLoss") or 0),
                    cb, p.get("currentDayProfitLoss"), p.get("currentDayProfitLossPercentage")
                ] + [None] * 19)
        wb.save(CONFIG["XLSX_FILE"])
        print("📂 Phase1 complete")

    def phase2(self, t: str):
        wb = load_workbook(CONFIG["XLSX_FILE"])
        ws = wb.active
        if ws is None:
            raise RuntimeError("Workbook.active returned None")
        ws = cast(Worksheet, ws)

        hdr = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
        rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, hdr["assetType"]).value == "OPTION"]
        f, u = set(), OptionFetcher(t)

        for r in rows:
            s = ws.cell(r, hdr["symbol"]).value
            if isinstance(s, str) and (p := parse(s)):
                f.add(p[0])

        # --- FIXED SECTION: Retrieve fundamental fields per symbol ---
        def _symbol_for_quotes(sym: str) -> str:
            sym = MAP.get(sym, sym)
            sym = sym.replace('/', '.').replace('-', '.')
            return sym

        fundamentals: Dict[str, Dict[str, Any]] = {}
        for sym in f:
            path_sym = _symbol_for_quotes(sym)
            url = f"https://api.schwabapi.com/marketdata/v1/{path_sym}/quotes"
            data = u.get(url, {"fields": "fundamental"})
            if not isinstance(data, dict):
                continue
            base_obj = data.get("quotes") if isinstance(data.get("quotes"), dict) else data
            if isinstance(base_obj, dict):
                keys_to_try = [sym, sym.upper(), path_sym, path_sym.upper(), sym.replace('/', '.'), sym.replace('-', '.')]
                for k in keys_to_try:
                    q = base_obj.get(k)
                    if isinstance(q, dict):
                        fnd = q.get("fundamental") or {}
                        fundamentals[sym] = {
                            "divYield": first(fnd.get("divYield"), fnd.get("dividendYield")),
                            "divAmount": fnd.get("divAmount"),
                            "divExDate": fnd.get("divExDate"),
                            "lastEarningsDate": fnd.get("lastEarningsDate"),
                            "nextDivExDate": fnd.get("nextDivExDate")
                        }
                        break

        ok, fail, start = 0, [], time.time()
        GREEN, RED, RESET = "\033[92m", "\033[91m", "\033[0m"

        for i, r in enumerate(rows, 1):
            s = ws.cell(r, hdr["symbol"]).value
            if not isinstance(s, str) or not (p := parse(s)):
                continue
            u_, e, cp, k = p
            print(f"[{i:>3}/{len(rows)}] {u_:<8} {e} {cp} {k:<7.2f} ⏱ {time.time()-start:6.1f}s ... ", end="", flush=True)
            d = u.extract(u.chain(u_, e, cp, k), cp, e, k, fundamentals.get(u_))
            if d:
                [ws.cell(r, hdr[k2], v) for k2, v in d.items() if k2 in hdr]
                ok += 1
                print(f"{GREEN}✅ updated{RESET}")
            else:
                fail.append(s)
                print(f"{RED}❌ failed{RESET}")
        wb.save(CONFIG["XLSX_FILE"])
        print(f"\n📊 Done: {GREEN}{ok} updated{RESET}, {RED}{len(fail)} failed{RESET}")
        print(f"⏳ Total elapsed: {time.time()-start:.1f}s")

# --------------------------------------------------------------------------
# Entry
# --------------------------------------------------------------------------
def main():
    t = SchwabAuth().refresh()
    r = requests.get(CONFIG["ACCT_URL"], headers={"Authorization": f"Bearer {t}"}, params={"fields": "positions"})
    r.raise_for_status()
    d = r.json()
    Path(CONFIG["RAW_FILE"]).write_text(json.dumps(d, indent=2))
    a = d if isinstance(d, list) else [d] if "securitiesAccount" in d else d.get("accounts", [])
    if not a:
        return print("⚠️ No accounts found")
    e = Exporter()
    e.phase1(a)
    e.phase2(t)

if __name__ == "__main__":
    main()
